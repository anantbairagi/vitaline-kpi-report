"""Clinical KPI Monthly Report Generator

Generates an Excel report with 4 clinical KPIs for Vitaline patients,
with all intermediate columns visible for auditability.

KPI 1: Falls with Major Injury (LS) — mirrors CMS N013.02
KPI 2: Prevalence of Falls (LS) — mirrors CMS N032.02
KPI 3: Pre/Post 90-Day Falls Comparison
KPI 4: Pre/Post 90-Day Hospitalization Comparison

Source: CMS MDS 3.0 Quality Measures User's Manual V17
        Effective January 1, 2025
Usage:  python scripts/generate_kpi_report.py [--year 2026] [--month 1]
"""

import argparse
import calendar
import sqlite3
from datetime import date, timedelta
from pathlib import Path

import pandas as pd

# ── Configuration ────────────────────────────────────────────────────────────

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "deid_clinical.db"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "reports"

OBRA_RFAS = ("01", "02", "03", "04", "05", "06")
PPS_RFAS = ("01", "02", "03", "04", "05", "06")
DISCHARGE_RFAS = ("10", "11")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _period_key(dt):
    return f"{dt.year}-{dt.month:02d}"


def _month_offset(ym_str, n):
    y, m = int(ym_str[:4]), int(ym_str[5:])
    m += n
    while m < 1:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    return f"{y}-{m:02d}"


def _qualifying_mask(df):
    """Per CMS v17 Ch1 S4: A0310A in [01-06] OR A0310B in [01-06] OR A0310F in [10,11]."""
    return (
        df["assessment_type_obra"].isin(OBRA_RFAS)
        | df["assessment_type_pps"].isin(PPS_RFAS)
        | df["entry_discharge_reporting"].isin(DISCHARGE_RFAS)
    )


def _rate(num, denom):
    return num / denom if denom > 0 else None


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_data():
    conn = sqlite3.connect(DB_PATH)
    mds = pd.read_sql("SELECT * FROM mds_clinical", conn)
    vitaline = pd.read_sql("SELECT * FROM vitaline", conn)
    facilities = pd.read_sql("SELECT * FROM facility_lookup", conn)
    conn.close()

    for col in ("entry_date", "discharge_date", "admission_date", "assessment_reference_date"):
        mds[col] = pd.to_datetime(mds[col], errors="coerce")

    for col in ("assessment_type_pps", "entry_type", "interrupted_stay"):
        if col not in mds.columns:
            mds[col] = None

    mds["assessment_type_pps"] = mds["assessment_type_pps"].fillna("99")

    # CMS v17 Ch1 S1: target_date per record type
    def _target_date(row):
        if row["entry_discharge_reporting"] == "01":
            return row["entry_date"]
        elif row["entry_discharge_reporting"] in ("10", "11", "12"):
            return row["discharge_date"]
        else:
            return row["assessment_reference_date"]

    mds["target_date"] = mds.apply(_target_date, axis=1)

    vitaline["clinic_date"] = pd.to_datetime(vitaline["clinic_date"], errors="coerce")
    return mds, vitaline, facilities


# ── Step 1: Eligible Patients ────────────────────────────────────────────────

def step1_eligible(vitaline, facilities, report_ym):
    """Identify Vitaline patients eligible for the report month.

    Criteria: participation='Received', visited in report month,
    3+ contiguous months of visits (1-month gap allowed).
    """
    v = vitaline[vitaline["participation"] == "Received"].copy()
    v["ym"] = v["clinic_date"].apply(_period_key)

    active = v[v["ym"] == report_ym]
    active_pids = active["surrogate_patient_id"].unique()

    pid_months = v.groupby("surrogate_patient_id")["ym"].apply(set).to_dict()
    pid_facility = (
        v.sort_values("clinic_date")
        .drop_duplicates("surrogate_patient_id", keep="last")[
            ["surrogate_patient_id", "surrogate_facility_id"]
        ]
        .set_index("surrogate_patient_id")["surrogate_facility_id"]
        .to_dict()
    )

    rows = []
    for pid in active_pids:
        months_set = pid_months.get(pid, set())
        chain = [report_ym]
        current = report_ym
        while True:
            m1 = _month_offset(current, -1)
            m2 = _month_offset(current, -2)
            if m1 in months_set:
                chain.append(m1)
                current = m1
            elif m2 in months_set:
                chain.append(m2)
                current = m2
            else:
                break

        chain_sorted = sorted(chain)
        rows.append({
            "surrogate_patient_id": pid,
            "surrogate_facility_id": pid_facility.get(pid, ""),
            "all_visit_months": ", ".join(sorted(months_set)),
            "contiguous_chain": ", ".join(chain_sorted),
            "chain_length": len(chain),
            "eligible": len(chain) >= 3,
        })

    df = pd.DataFrame(rows)
    df = df.merge(facilities, on="surrogate_facility_id", how="left")
    return df


# ── Step 2: Long-Stay (CDIF) ────────────────────────────────────────────────

def _build_stays_and_cdif(patient_mds, target_period_end):
    """Build stays/episodes and compute CDIF for one patient-facility pair.

    Implements CMS QM User's Manual V17, Chapter 1 Section 1 & Chapter 4.
    Uses A1700 (entry_type) directly for admission/reentry classification.
    Returns (episode_start, episode_end, cdif, num_stays, detail_str).
    """
    events = []
    for _, row in patient_mds.iterrows():
        atype = row["entry_discharge_reporting"]
        if atype == "01" and pd.notna(row["entry_date"]):
            events.append(("E", row["entry_date"], atype, row.get("entry_type")))
        elif atype in ("10", "11", "12") and pd.notna(row["discharge_date"]):
            events.append(("D", row["discharge_date"], atype, None))

    if not events:
        adm = patient_mds["admission_date"].dropna().min()
        if pd.notna(adm):
            cdif = (target_period_end - adm).days + 1
            return adm, target_period_end, cdif, 1, \
                f"Fallback: {adm.date()} -> {target_period_end.date()} ({cdif}d)"
        return None, None, 0, 0, "No entry/admission records"

    events.sort(key=lambda e: (e[1], 0 if e[0] == "E" else 1))

    stays = []
    open_entry = None
    open_entry_type = None
    for ev in events:
        if ev[0] == "E":
            if open_entry is not None:
                imputed_end = ev[1] - timedelta(days=1)
                if imputed_end >= open_entry:
                    stays.append({"start": open_entry, "end": imputed_end,
                                  "dt": None, "etype": open_entry_type})
            open_entry = ev[1]
            open_entry_type = ev[3]
        else:
            if open_entry is not None:
                stays.append({"start": open_entry, "end": ev[1],
                              "dt": ev[2], "etype": open_entry_type})
                open_entry = None
                open_entry_type = None

    if open_entry is not None:
        stays.append({"start": open_entry, "end": target_period_end,
                      "dt": None, "etype": open_entry_type})

    if not stays:
        return None, None, 0, 0, "No valid stays"

    stays.sort(key=lambda s: s["start"])

    # Build latest episode using A1700 directly (CMS v17 Ch4)
    episode = [stays[-1]]
    for k in range(len(stays) - 2, -1, -1):
        earliest = episode[0]
        if earliest.get("etype") == "2":
            episode.insert(0, stays[k])
            continue
        elif earliest.get("etype") == "1":
            break
        else:
            prev = stays[k]
            if prev["dt"] == "11":
                gap = (earliest["start"] - prev["end"]).days
                if gap <= 30:
                    episode.insert(0, prev)
                    continue
            break

    ep_start = episode[0]["start"]
    latest_stay = episode[-1]
    ep_end = latest_stay["end"]

    cdif = 0
    details = []
    for s in episode:
        ongoing = s["dt"] is None
        if ongoing:
            days = (s["end"] - s["start"]).days + 1
        else:
            days = max(1, (s["end"] - s["start"]).days)
        cdif += days
        tag = " ongoing" if ongoing and s["end"] == target_period_end else ""
        details.append(f"{s['start'].date()}->{s['end'].date()} ({days}d{tag})")

    return ep_start, ep_end, cdif, len(episode), "; ".join(details)


def step2_longstay(mds, eligible_df, target_period_end):
    """Compute CDIF and long-stay classification for eligible patients."""
    eligible = eligible_df[eligible_df["eligible"]]
    rows = []
    for _, pat in eligible.iterrows():
        pid = pat["surrogate_patient_id"]
        fid = pat["surrogate_facility_id"]
        pm = mds[(mds["surrogate_patient_id"] == pid) & (mds["surrogate_facility_id"] == fid)]

        ep_start, ep_end, cdif, n_stays, detail = _build_stays_and_cdif(pm, target_period_end)
        rows.append({
            "surrogate_patient_id": pid,
            "surrogate_facility_id": fid,
            "episode_start": ep_start.date() if pd.notna(ep_start) else None,
            "episode_end": ep_end.date() if ep_end is not None and pd.notna(ep_end) else None,
            "num_stays": n_stays,
            "stay_details": detail,
            "cdif": cdif,
            "is_long_stay": cdif >= 101,
        })

    return pd.DataFrame(rows)


# ── Step 3-4: KPI 1 & KPI 2 ─────────────────────────────────────────────────

def _assess_quality_measure(patient_mds, item_col, numerator_values,
                            episode_start=None, episode_end=None,
                            cms_lookback=False,
                            window_start=None, window_end=None):
    """Evaluate a CMS quality measure for one patient.

    Monthly (cms_lookback=False): assessments with target_date in
    [window_start, window_end].

    CMS mirror (cms_lookback=True): per V17 Ch1 S4 —
    1. Target assessment = latest qualifying RFA within episode,
       target_date no more than 120 days before episode_end.
    2. Look-back scan = qualifying RFAs within episode,
       target_date no more than 275 days before target assessment.
    """
    qualifying = patient_mds[_qualifying_mask(patient_mds)].copy()

    if episode_start is not None and pd.notna(episode_start):
        qualifying = qualifying[qualifying["target_date"] >= pd.Timestamp(episode_start)]

    if cms_lookback:
        if episode_end is None:
            return _empty_result("No episode_end for CMS lookback")

        ep_end_ts = pd.Timestamp(episode_end)
        floor_120 = ep_end_ts - timedelta(days=120)

        candidates = qualifying[
            (qualifying["target_date"] >= floor_120)
            & (qualifying["target_date"] <= ep_end_ts)
        ]

        if candidates.empty:
            return _empty_result("No qualifying assessment within 120 days of episode end")

        target_td = candidates["target_date"].max()
        lookback_floor = target_td - timedelta(days=275)
        if episode_start is not None and pd.notna(episode_start):
            lookback_floor = max(lookback_floor, pd.Timestamp(episode_start))

        scan = qualifying[
            (qualifying["target_date"] <= target_td)
            & (qualifying["target_date"] >= lookback_floor)
        ]

        scan_label = (f"Target: {target_td.date()}, scan: "
                      f"{lookback_floor.date()} to {target_td.date()} "
                      f"({len(scan)} assessments)")
    else:
        scan = qualifying[
            (qualifying["target_date"] >= window_start)
            & (qualifying["target_date"] <= window_end)
        ]
        scan_label = ""

    if scan.empty:
        return _empty_result("No assessments in scan window")

    values = scan[item_col].tolist()
    coded = [v for v in values if pd.notna(v) and v not in ("-", "", "^")]
    excluded = len(coded) == 0
    in_num = any(v in numerator_values for v in coded) if coded else False

    return {
        "has_assessment": True,
        "excluded": excluded,
        "in_numerator": in_num,
        "n_assessments": len(scan),
        "item_values": ", ".join(str(v) for v in values),
        "scan_detail": scan_label,
    }


def _empty_result(reason):
    return {
        "has_assessment": False, "excluded": False, "in_numerator": False,
        "n_assessments": 0, "item_values": "", "scan_detail": reason,
    }


def step3_kpi1(mds, eligible_df, longstay_df):
    """KPI 1: Falls with Major Injury — CMS N013.02 exact mirror."""
    ls = longstay_df[longstay_df["is_long_stay"]]
    ls_pids = set(ls["surrogate_patient_id"])

    rows = []
    for _, pat in eligible_df[eligible_df["eligible"]].iterrows():
        pid = pat["surrogate_patient_id"]
        fid = pat["surrogate_facility_id"]
        is_ls = pid in ls_pids

        if not is_ls:
            rows.append({
                "surrogate_patient_id": pid, "surrogate_facility_id": fid,
                "is_long_stay": False,
                "has_assessment": None, "excluded": None,
                "in_numerator": None, "n_assessments": 0,
                "item_values": "", "scan_detail": "",
            })
            continue

        ep_row = ls[ls["surrogate_patient_id"] == pid].iloc[0]
        ep_start = pd.Timestamp(ep_row["episode_start"]) if ep_row["episode_start"] else None
        ep_end = pd.Timestamp(ep_row["episode_end"]) if ep_row["episode_end"] else None
        pm = mds[(mds["surrogate_patient_id"] == pid) & (mds["surrogate_facility_id"] == fid)]

        r = _assess_quality_measure(pm, "j1900c_major_injury", ("1", "2"),
                                    episode_start=ep_start, episode_end=ep_end,
                                    cms_lookback=True)

        rows.append({
            "surrogate_patient_id": pid, "surrogate_facility_id": fid,
            "is_long_stay": True, **r,
        })

    return pd.DataFrame(rows)


def step4_kpi2(mds, eligible_df, longstay_df):
    """KPI 2: Prevalence of Falls — CMS N032.02 exact mirror."""
    ls = longstay_df[longstay_df["is_long_stay"]]
    ls_pids = set(ls["surrogate_patient_id"])

    rows = []
    for _, pat in eligible_df[eligible_df["eligible"]].iterrows():
        pid = pat["surrogate_patient_id"]
        fid = pat["surrogate_facility_id"]
        is_ls = pid in ls_pids

        if not is_ls:
            rows.append({
                "surrogate_patient_id": pid, "surrogate_facility_id": fid,
                "is_long_stay": False,
                "has_assessment": None, "excluded": None,
                "in_numerator": None, "n_assessments": 0,
                "item_values": "", "scan_detail": "",
            })
            continue

        ep_row = ls[ls["surrogate_patient_id"] == pid].iloc[0]
        ep_start = pd.Timestamp(ep_row["episode_start"]) if ep_row["episode_start"] else None
        ep_end = pd.Timestamp(ep_row["episode_end"]) if ep_row["episode_end"] else None
        pm = mds[(mds["surrogate_patient_id"] == pid) & (mds["surrogate_facility_id"] == fid)]

        r = _assess_quality_measure(pm, "j1800_any_fall", ("1",),
                                    episode_start=ep_start, episode_end=ep_end,
                                    cms_lookback=True)

        rows.append({
            "surrogate_patient_id": pid, "surrogate_facility_id": fid,
            "is_long_stay": True, **r,
        })

    return pd.DataFrame(rows)


# ── Step 5: KPI 3 — Pre/Post Falls ──────────────────────────────────────────

def _find_pre_90_window(clinic_dates_sorted):
    """Find the most recent 90-day window without any Vitaline infusion."""
    if not clinic_dates_sorted:
        return None, None

    for i in range(len(clinic_dates_sorted) - 1, 0, -1):
        gap = (clinic_dates_sorted[i] - clinic_dates_sorted[i - 1]).days
        if gap > 90:
            pre_end = clinic_dates_sorted[i] - timedelta(days=1)
            pre_start = clinic_dates_sorted[i] - timedelta(days=90)
            return pre_start, pre_end

    pre_end = clinic_dates_sorted[0] - timedelta(days=1)
    pre_start = clinic_dates_sorted[0] - timedelta(days=90)
    return pre_start, pre_end


def _count_falls_in_window(patient_mds, window_start, window_end):
    """Count J1900C values from qualifying assessments within a window."""
    qual = patient_mds[_qualifying_mask(patient_mds)]
    in_window = qual[
        (qual["target_date"] >= window_start)
        & (qual["target_date"] <= window_end)
    ]

    n_assess = len(in_window)
    j1900c_vals = in_window["j1900c_major_injury"].dropna().tolist()
    coded = [v for v in j1900c_vals if v not in ("-", "", "^")]

    count_1 = sum(1 for v in coded if v == "1")
    count_2 = sum(1 for v in coded if v == "2")

    return {
        "n_assessments": n_assess,
        "j1900c_values": ", ".join(str(v) for v in j1900c_vals) if j1900c_vals else "",
        "j1900c_eq1_count": count_1,
        "j1900c_eq2_count": count_2,
        "has_any_fall": (count_1 + count_2) > 0,
        "has_j1900c_2": count_2 > 0,
    }


def step5_kpi3(mds, vitaline, eligible_df, post_start, post_end):
    """KPI 3: Pre/Post 90-day falls comparison for all eligible patients."""
    v_received = vitaline[vitaline["participation"] == "Received"]

    rows = []
    for _, pat in eligible_df[eligible_df["eligible"]].iterrows():
        pid = pat["surrogate_patient_id"]
        fid = pat["surrogate_facility_id"]

        clinic_dates = sorted(
            v_received[v_received["surrogate_patient_id"] == pid]["clinic_date"]
            .dropna().dt.date.tolist()
        )
        pre_start, pre_end = _find_pre_90_window(clinic_dates)

        pm = mds[(mds["surrogate_patient_id"] == pid) & (mds["surrogate_facility_id"] == fid)]

        empty_pre = {"n_assessments": 0, "j1900c_values": "", "j1900c_eq1_count": 0,
                     "j1900c_eq2_count": 0, "has_any_fall": False, "has_j1900c_2": False}
        pre = _count_falls_in_window(pm, pd.Timestamp(pre_start), pd.Timestamp(pre_end)) \
            if pre_start else empty_pre
        post = _count_falls_in_window(pm, post_start, post_end)

        gap_source = "pre-first-visit"
        if clinic_dates and len(clinic_dates) > 1:
            for i in range(len(clinic_dates) - 1, 0, -1):
                if (clinic_dates[i] - clinic_dates[i - 1]).days > 90:
                    gap_source = f"gap between {clinic_dates[i-1]} and {clinic_dates[i]}"
                    break

        rows.append({
            "surrogate_patient_id": pid, "surrogate_facility_id": fid,
            "first_vitaline_date": str(clinic_dates[0]) if clinic_dates else "",
            "pre_window_source": gap_source,
            "pre_window_start": str(pre_start) if pre_start else "",
            "pre_window_end": str(pre_end) if pre_end else "",
            "post_window_start": str(post_start.date()),
            "post_window_end": str(post_end.date()),
            "pre_n_assessments": pre["n_assessments"],
            "pre_j1900c_values": pre["j1900c_values"],
            "pre_j1900c_eq1": pre["j1900c_eq1_count"],
            "pre_j1900c_eq2": pre["j1900c_eq2_count"],
            "pre_any_fall": pre["has_any_fall"],
            "pre_has_j1900c_2": pre["has_j1900c_2"],
            "post_n_assessments": post["n_assessments"],
            "post_j1900c_values": post["j1900c_values"],
            "post_j1900c_eq1": post["j1900c_eq1_count"],
            "post_j1900c_eq2": post["j1900c_eq2_count"],
            "post_any_fall": post["has_any_fall"],
            "post_has_j1900c_2": post["has_j1900c_2"],
        })

    return pd.DataFrame(rows)


# ── Step 6: KPI 4 — Pre/Post Hospitalizations ───────────────────────────────

def _count_hosp_in_window(patient_mds, window_start, window_end):
    """Count hospitalization events (A0310F in 10,11 + discharge_status=04)."""
    discharges = patient_mds[
        patient_mds["entry_discharge_reporting"].isin(("10", "11"))
    ]
    in_window = discharges[
        (discharges["target_date"] >= window_start)
        & (discharges["target_date"] <= window_end)
    ]

    to_hospital = in_window[in_window["discharge_status"] == "04"]
    n_hosp = len(to_hospital)

    return {
        "n_discharges": len(in_window),
        "n_hospitalizations": n_hosp,
        "discharge_statuses": ", ".join(in_window["discharge_status"].dropna().tolist()),
        "has_1plus": n_hosp >= 1,
        "has_2plus": n_hosp >= 2,
    }


def step6_kpi4(mds, vitaline, eligible_df, post_start, post_end):
    """KPI 4: Pre/Post 90-day hospitalization comparison."""
    v_received = vitaline[vitaline["participation"] == "Received"]

    rows = []
    for _, pat in eligible_df[eligible_df["eligible"]].iterrows():
        pid = pat["surrogate_patient_id"]
        fid = pat["surrogate_facility_id"]

        clinic_dates = sorted(
            v_received[v_received["surrogate_patient_id"] == pid]["clinic_date"]
            .dropna().dt.date.tolist()
        )
        pre_start, pre_end = _find_pre_90_window(clinic_dates)

        pm = mds[(mds["surrogate_patient_id"] == pid) & (mds["surrogate_facility_id"] == fid)]

        empty_pre = {"n_discharges": 0, "n_hospitalizations": 0,
                     "discharge_statuses": "", "has_1plus": False, "has_2plus": False}
        pre = _count_hosp_in_window(pm, pd.Timestamp(pre_start), pd.Timestamp(pre_end)) \
            if pre_start else empty_pre
        post = _count_hosp_in_window(pm, post_start, post_end)

        gap_source = "pre-first-visit"
        if clinic_dates and len(clinic_dates) > 1:
            for i in range(len(clinic_dates) - 1, 0, -1):
                if (clinic_dates[i] - clinic_dates[i - 1]).days > 90:
                    gap_source = f"gap between {clinic_dates[i-1]} and {clinic_dates[i]}"
                    break

        rows.append({
            "surrogate_patient_id": pid, "surrogate_facility_id": fid,
            "pre_window_start": str(pre_start) if pre_start else "",
            "pre_window_end": str(pre_end) if pre_end else "",
            "post_window_start": str(post_start.date()),
            "post_window_end": str(post_end.date()),
            "pre_window_source": gap_source,
            "pre_n_discharges": pre["n_discharges"],
            "pre_n_hospitalizations": pre["n_hospitalizations"],
            "pre_discharge_statuses": pre["discharge_statuses"],
            "pre_has_1plus_hosp": pre["has_1plus"],
            "pre_has_2plus_hosp": pre["has_2plus"],
            "post_n_discharges": post["n_discharges"],
            "post_n_hospitalizations": post["n_hospitalizations"],
            "post_discharge_statuses": post["discharge_statuses"],
            "post_has_1plus_hosp": post["has_1plus"],
            "post_has_2plus_hosp": post["has_2plus"],
        })

    return pd.DataFrame(rows)


# ── Summary ──────────────────────────────────────────────────────────────────

def compute_summary(eligible_df, longstay_df, kpi1_df, kpi2_df, kpi3_df, kpi4_df,
                    facilities, report_ym):
    """Build per-facility summary of all KPIs."""
    elig = eligible_df[eligible_df["eligible"]]
    facility_ids = sorted(elig["surrogate_facility_id"].unique())

    summary_rows = []
    for fid in [None] + list(facility_ids):
        if fid is None:
            label, company = "OVERALL", ""
            e, ls, k1, k2, k3, k4 = elig, longstay_df, kpi1_df, kpi2_df, kpi3_df, kpi4_df
        else:
            label = fid
            co = facilities.loc[facilities["surrogate_facility_id"] == fid, "company"].values
            company = co[0] if len(co) else ""
            e = elig[elig["surrogate_facility_id"] == fid]
            ls = longstay_df[longstay_df["surrogate_facility_id"] == fid]
            k1 = kpi1_df[kpi1_df["surrogate_facility_id"] == fid]
            k2 = kpi2_df[kpi2_df["surrogate_facility_id"] == fid]
            k3 = kpi3_df[kpi3_df["surrogate_facility_id"] == fid]
            k4 = kpi4_df[kpi4_df["surrogate_facility_id"] == fid]

        n_eligible = len(e)
        n_longstay = int(ls["is_long_stay"].sum())

        def _kpi_stats(kdf):
            kls = kdf[kdf["is_long_stay"] == True]
            ha = kls["has_assessment"]
            ex = kls["excluded"]
            inn = kls["in_numerator"]
            d = int(((ha == True) & (ex == False)).sum())
            n = int((inn == True).sum())
            return d, n, _rate(n, d)

        k1_d, k1_n, k1_r = _kpi_stats(k1)
        k2_d, k2_n, k2_r = _kpi_stats(k2)

        n3 = len(k3)
        k3pa = int(k3["pre_any_fall"].sum()) if n3 else 0
        k3oa = int(k3["post_any_fall"].sum()) if n3 else 0
        k3pb = int(k3["pre_has_j1900c_2"].sum()) if n3 else 0
        k3ob = int(k3["post_has_j1900c_2"].sum()) if n3 else 0

        n4 = len(k4)
        k4pa = int(k4["pre_has_1plus_hosp"].sum()) if n4 else 0
        k4oa = int(k4["post_has_1plus_hosp"].sum()) if n4 else 0
        k4pb = int(k4["pre_has_2plus_hosp"].sum()) if n4 else 0
        k4ob = int(k4["post_has_2plus_hosp"].sum()) if n4 else 0

        summary_rows.append({
            "Facility": label, "Company": company,
            "Eligible Patients": n_eligible, "Long-Stay Patients": n_longstay,
            "KPI1 Denom": k1_d, "KPI1 Num": k1_n, "KPI1 Rate": k1_r,
            "KPI2 Denom": k2_d, "KPI2 Num": k2_n, "KPI2 Rate": k2_r,
            "KPI3 Eligible": n3,
            "KPI3-A Pre (>=1 fall)": k3pa, "KPI3-A Post (>=1 fall)": k3oa,
            "KPI3-A Pre %": _rate(k3pa, n3), "KPI3-A Post %": _rate(k3oa, n3),
            "KPI3-B Pre (J1900C=2)": k3pb, "KPI3-B Post (J1900C=2)": k3ob,
            "KPI3-B Pre %": _rate(k3pb, n3), "KPI3-B Post %": _rate(k3ob, n3),
            "KPI4 Eligible": n4,
            "KPI4-A Pre (>=1 hosp)": k4pa, "KPI4-A Post (>=1 hosp)": k4oa,
            "KPI4-A Pre %": _rate(k4pa, n4), "KPI4-A Post %": _rate(k4oa, n4),
            "KPI4-B Pre (>=2 hosp)": k4pb, "KPI4-B Post (>=2 hosp)": k4ob,
            "KPI4-B Pre %": _rate(k4pb, n4), "KPI4-B Post %": _rate(k4ob, n4),
        })

    return pd.DataFrame(summary_rows)


# ── Excel Generation ─────────────────────────────────────────────────────────

def _definitions_rows():
    return [
        ("Sheet", "Column / Metric", "Definition"),
        ("", "", ""),
        ("General", "Report Month", "The calendar month for which KPIs are computed"),
        ("General", "Eligible Patient", "Vitaline patient with participation='Received', "
         "visited in report month, 3+ contiguous months (1-month gap allowed)"),
        ("General", "Long-Stay (CMS)", "CDIF >= 101 days per CMS QM User's Manual V17. "
         "Episode constructed using A1700 for admission/reentry. "
         "Include entry day (A1600), exclude discharge day (A2000)."),
        ("General", "Target Date", "Per CMS V17: entry_date for A0310F=01, "
         "discharge_date for A0310F=10/11/12, ARD (A2300) for all others."),
        ("General", "Qualifying RFA", "A0310A in [01-06] OR A0310B in [01-06] OR A0310F in [10,11]"),
        ("", "", ""),
        ("KPI 1", "Falls with Major Injury (LS)", "CMS N013.02 mirror. "
         "% of long-stay residents with J1900C = [1,2] on any look-back scan assessment."),
        ("KPI 1", "CMS Quarterly", "Target assessment: latest qualifying within 120 days "
         "of episode end. Look-back scan: 275 days back from target within episode. "
         "Covers ~1 year of fall history."),
        ("KPI 1", "Monthly Snapshot", "Assessments with target_date in report month only."),
        ("KPI 1", "Denominator", "Long-stay patients with 1+ look-back scan assessments "
         "where J1900C is coded (not NULL, dash, or caret)."),
        ("KPI 1", "Numerator", "Patients in denominator with J1900C in (1,2)."),
        ("KPI 1", "Exclusion", "J1900C not coded on ALL look-back scan assessments."),
        ("", "", ""),
        ("KPI 2", "Prevalence of Falls (LS)", "CMS N032.02 mirror. "
         "% of long-stay residents with J1800 = [1] on any look-back scan assessment."),
        ("KPI 2", "Denominator", "Same as KPI 1 but checking J1800 coding."),
        ("KPI 2", "Numerator", "J1800 = 1 on at least one scan assessment."),
        ("", "", ""),
        ("KPI 3", "Pre/Post 90-Day Falls", "Compare major-injury falls before/after Vitaline."),
        ("KPI 3", "PRE Window", "Most recent 90-day gap without Vitaline infusions."),
        ("KPI 3", "POST Window", "90 days ending on last day of report month."),
        ("KPI 3", "Part A", "% with at least 1 assessment where J1900C in (1,2)."),
        ("KPI 3", "Part B", "% with at least 1 assessment where J1900C = 2."),
        ("", "", ""),
        ("KPI 4", "Pre/Post Hospitalizations", "Compare hospitalizations before/after Vitaline."),
        ("KPI 4", "Hospitalization", "A0310F in [10,11] + discharge_status = '04'."),
        ("", "", ""),
        ("Sources", "CMS QM Manual V17", "MDS 3.0 Quality Measures User's Manual V17, "
         "Effective January 1, 2025"),
        ("Sources", "CDIF / Episode", "V17 Chapter 1 S1-S4, Chapter 4"),
        ("Sources", "Falls Measures", "V17 Table 2-12 (N013.02), Table 2-32 (N032.02)"),
    ]


def write_excel(summary_df, eligible_df, longstay_df, kpi1_df, kpi2_df,
                kpi3_df, kpi4_df, report_year, report_month):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OUTPUT_DIR.mkdir(exist_ok=True)
    fname = OUTPUT_DIR / f"clinical_kpi_report_{report_year}_{report_month:02d}.xlsx"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_sheet(writer, df, sheet_name):
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        ws = writer.sheets[sheet_name]
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(
                len(str(df.columns[col_idx - 1])),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                    cell.number_format = "0.00%"

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        _write_sheet(writer, summary_df, "Summary")
        _write_sheet(writer, eligible_df, "1_Eligible_Patients")
        _write_sheet(writer, longstay_df, "2_Long_Stay")
        _write_sheet(writer, kpi1_df, "3_KPI1_Falls_MajorInjury")
        _write_sheet(writer, kpi2_df, "4_KPI2_Prevalence_Falls")
        _write_sheet(writer, kpi3_df, "5_KPI3_PrePost_Falls")
        _write_sheet(writer, kpi4_df, "6_KPI4_PrePost_Hosp")
        defs = pd.DataFrame(_definitions_rows()[1:], columns=_definitions_rows()[0])
        _write_sheet(writer, defs, "7_Definitions")

    print(f"Report written to {fname}")
    return fname


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Clinical KPI Monthly Report")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--month", type=int, default=1)
    args = parser.parse_args()

    report_year, report_month = args.year, args.month
    report_ym = f"{report_year}-{report_month:02d}"
    report_start = pd.Timestamp(date(report_year, report_month, 1))
    report_end = pd.Timestamp(date(
        report_year, report_month, calendar.monthrange(report_year, report_month)[1]))
    quarterly_start = report_start - pd.DateOffset(months=2)
    quarterly_start = pd.Timestamp(date(quarterly_start.year, quarterly_start.month, 1))
    post_90_end = report_end
    post_90_start = post_90_end - timedelta(days=89)

    print(f"Generating Clinical KPI Report for {calendar.month_name[report_month]} {report_year}")
    print(f"  Report month:     {report_start.date()} to {report_end.date()}")
    print(f"  Quarterly window: {quarterly_start.date()} to {report_end.date()}")
    print(f"  POST 90-day:      {post_90_start.date()} to {post_90_end.date()}")
    print(f"  Source: CMS QM User's Manual V17 (Jan 2025)")
    print()

    print("Loading data...")
    mds, vitaline, facilities = load_data()
    print(f"  MDS records:      {len(mds):,}")
    print(f"  Vitaline records: {len(vitaline):,}")
    print(f"  Facilities:       {len(facilities)}")
    print()

    print("Step 1: Identifying eligible patients...")
    eligible_df = step1_eligible(vitaline, facilities, report_ym)
    n_active = len(eligible_df)
    n_eligible = int(eligible_df["eligible"].sum())
    print(f"  Active in {report_ym}: {n_active}")
    print(f"  Eligible (3+ months): {n_eligible}")
    print()

    print("Step 2: Computing long-stay (CDIF with A1700)...")
    longstay_df = step2_longstay(mds, eligible_df, report_end)
    n_ls = int(longstay_df["is_long_stay"].sum())
    print(f"  Long-stay (CDIF >= 101): {n_ls} / {n_eligible}")
    print()

    print("Step 3: KPI 1 -- Falls with Major Injury (N013.02)...")
    kpi1_df = step3_kpi1(mds, eligible_df, longstay_df)
    print("  Done.")

    print("Step 4: KPI 2 -- Prevalence of Falls (N032.02)...")
    kpi2_df = step4_kpi2(mds, eligible_df, longstay_df)
    print("  Done.")

    print("Step 5: KPI 3 -- Pre/Post 90-day Falls...")
    kpi3_df = step5_kpi3(mds, vitaline, eligible_df, post_90_start, post_90_end)
    print("  Done.")

    print("Step 6: KPI 4 -- Pre/Post 90-day Hospitalizations...")
    kpi4_df = step6_kpi4(mds, vitaline, eligible_df, post_90_start, post_90_end)
    print("  Done.")
    print()

    print("Computing summary...")
    summary_df = compute_summary(eligible_df, longstay_df, kpi1_df, kpi2_df,
                                 kpi3_df, kpi4_df, facilities, report_ym)

    overall = summary_df[summary_df["Facility"] == "OVERALL"].iloc[0]
    print()
    print("=" * 60)
    print(f"  HEADLINE RESULTS -- {calendar.month_name[report_month]} {report_year}")
    print("=" * 60)
    print(f"  Eligible Patients:  {int(overall['Eligible Patients'])}")
    print(f"  Long-Stay:          {int(overall['Long-Stay Patients'])}")
    print()
    for label, prefix in [("KPI 1 (Falls Major Injury, N013.02)", "KPI1"),
                          ("KPI 2 (Prevalence of Falls, N032.02)", "KPI2")]:
        r = overall[f"{prefix} Rate"]
        print(f"  {label}:")
        if r is not None:
            print(f"    {int(overall[f'{prefix} Num'])}/{int(overall[f'{prefix} Denom'])} = {r:.1%}")
        else:
            print(f"    N/A")
    print()
    for label, prefix in [("KPI 3 (Pre/Post Falls)", "KPI3"), ("KPI 4 (Pre/Post Hosp)", "KPI4")]:
        pa = overall[f"{prefix}-A Pre %"]
        oa = overall[f"{prefix}-A Post %"]
        pb = overall[f"{prefix}-B Pre %"]
        ob = overall[f"{prefix}-B Post %"]
        da = ">=1 fall" if prefix == "KPI3" else ">=1 hosp"
        db = "J1900C=2" if prefix == "KPI3" else ">=2 hosp"
        print(f"  {label}:")
        if pa is not None and oa is not None:
            print(f"    Part A ({da}): Pre {pa:.1%} -> Post {oa:.1%}")
        if pb is not None and ob is not None:
            print(f"    Part B ({db}): Pre {pb:.1%} -> Post {ob:.1%}")
    print("=" * 60)
    print()

    print("Writing Excel...")
    outpath = write_excel(summary_df, eligible_df, longstay_df,
                          kpi1_df, kpi2_df, kpi3_df, kpi4_df,
                          report_year, report_month)
    print(f"\nDone! Report saved to: {outpath}")


if __name__ == "__main__":
    main()
